"""Shared report exporters.

PDF
---
Uses ReportLab's ``BaseDocTemplate`` with a custom ``PageTemplate`` so we get
a branded header (court name + optional logo) and a footer (generation
timestamp + page numbers) on every page.

Excel
-----
Uses openpyxl. Writes a title row, a metadata row (generated date + filters),
a blank row, then a styled header row and the data rows. Optionally writes
group sub-headers and subtotal rows for "grouped" reports.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, Sequence

from django.conf import settings
from django.http import HttpResponse


# ---------------------------------------------------------------------------
# Branding helpers
# ---------------------------------------------------------------------------

def _court_name() -> str:
    return getattr(settings, "COURT_NAME", "Customary Court of Appeal, FCT")


def _court_logo_path():
    """Absolute path to the court logo, or None if not configured / missing."""
    path = getattr(settings, "COURT_LOGO_PATH", None)
    if not path:
        return None
    from pathlib import Path
    p = Path(path)
    return str(p) if p.exists() else None


# ===========================================================================
# PDF
# ===========================================================================

def render_pdf(
    *,
    title: str,
    subtitle: str = "",
    filters_summary: str = "",
    columns: Sequence[str],
    rows: Sequence[Sequence],
    groups: Sequence[dict] | None = None,
    filename: str,
) -> HttpResponse:
    """Render a tabular report as a branded PDF.

    Parameters
    ----------
    title, subtitle : str
        Headline pieces shown above the table.
    filters_summary : str
        Free-form text describing the applied filters (e.g. "Department: HR;
        Location: Abuja"). Shown under the subtitle.
    columns : sequence of str
        Column headers.
    rows : sequence of sequences
        Flat-table rows. Ignored if ``groups`` is provided.
    groups : optional list of dicts
        For grouped reports. Each item: ``{"label": str, "rows": [...],
        "subtotal": int|None}``. When present, ``rows`` is ignored and the
        report is rendered as a sequence of grouped tables with subtotals.
    filename : str
        Suggested download filename.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        Image,
        PageBreak,
        PageTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    page_size = landscape(A4)
    page_w, page_h = page_size

    LEFT = RIGHT = 14 * mm
    TOP = 28 * mm  # leave room for header band
    BOTTOM = 16 * mm  # leave room for footer band

    doc = BaseDocTemplate(
        response,
        pagesize=page_size,
        leftMargin=LEFT,
        rightMargin=RIGHT,
        topMargin=TOP,
        bottomMargin=BOTTOM,
        title=title,
    )

    frame = Frame(
        LEFT, BOTTOM,
        page_w - LEFT - RIGHT,
        page_h - TOP - BOTTOM,
        showBoundary=0,
        id="content",
    )

    court_name = _court_name()
    logo_path = _court_logo_path()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    def _draw_chrome(canvas, _doc):
        canvas.saveState()
        # Header band
        canvas.setFillColor(colors.HexColor("#1f4e3d"))
        canvas.rect(0, page_h - 22 * mm, page_w, 22 * mm, stroke=0, fill=1)

        # Logo (optional)
        if logo_path:
            try:
                canvas.drawImage(
                    logo_path,
                    LEFT, page_h - 20 * mm,
                    width=16 * mm, height=16 * mm,
                    preserveAspectRatio=True, mask="auto",
                )
                text_x = LEFT + 20 * mm
            except Exception:
                text_x = LEFT
        else:
            text_x = LEFT

        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(text_x, page_h - 10 * mm, court_name)
        canvas.setFont("Helvetica", 9)
        canvas.drawString(text_x, page_h - 15 * mm, title)

        # Footer rule + text
        canvas.setStrokeColor(colors.HexColor("#cccccc"))
        canvas.setLineWidth(0.4)
        canvas.line(LEFT, 12 * mm, page_w - RIGHT, 12 * mm)
        canvas.setFillColor(colors.HexColor("#555555"))
        canvas.setFont("Helvetica", 8)
        canvas.drawString(LEFT, 7 * mm, f"Generated: {generated_at}")
        page_num = canvas.getPageNumber()
        canvas.drawRightString(
            page_w - RIGHT, 7 * mm, f"Page {page_num}"
        )
        canvas.drawCentredString(page_w / 2, 7 * mm, court_name)
        canvas.restoreState()

    doc.addPageTemplates([
        PageTemplate(id="branded", frames=[frame], onPage=_draw_chrome),
    ])

    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    h2 = ParagraphStyle(
        "h2", parent=styles["Heading2"], spaceAfter=2, fontSize=12
    )
    small = ParagraphStyle(
        "small", parent=body, fontSize=8, textColor=colors.HexColor("#555555")
    )

    story = []
    if subtitle:
        story.append(Paragraph(subtitle, h2))
    if filters_summary:
        story.append(Paragraph(f"<b>Filters:</b> {filters_summary}", small))
    story.append(Spacer(1, 6))

    base_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e3d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f4f6f9")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ])

    if groups:
        for idx, grp in enumerate(groups):
            if idx > 0:
                story.append(Spacer(1, 8))
            story.append(Paragraph(
                f'<b>{grp["label"]}</b> &nbsp;'
                f'<font color="#555555" size="8">({grp.get("subtotal", len(grp["rows"]))} records)</font>',
                h2,
            ))
            data = [list(columns)] + [list(r) for r in grp["rows"]]
            tbl = Table(data, repeatRows=1)
            tbl.setStyle(base_style)
            story.append(tbl)
        total = sum(grp.get("subtotal", len(grp["rows"])) for grp in groups)
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"<b>Grand total:</b> {total} records", body))
    else:
        data = [list(columns)] + [list(r) for r in rows]
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(base_style)
        story.append(tbl)
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"<b>Total:</b> {len(rows)} records", body))

    doc.build(story)
    return response


# ===========================================================================
# Excel
# ===========================================================================

def render_excel(
    *,
    title: str,
    subtitle: str = "",
    filters_summary: str = "",
    columns: Sequence[str],
    rows: Sequence[Sequence],
    groups: Sequence[dict] | None = None,
    filename: str,
    sheet_name: str = "Report",
) -> HttpResponse:
    """Render a tabular report as a styled .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Report"

    court = _court_name()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    last_col_letter = get_column_letter(max(1, len(columns)))

    header_fill = PatternFill("solid", fgColor="1F4E3D")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1F4E3D")
    sub_font = Font(italic=True, size=10, color="555555")
    group_fill = PatternFill("solid", fgColor="D6B87A")
    group_font = Font(bold=True, size=11, color="1F1F1F")
    subtotal_font = Font(bold=True, italic=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    ws.cell(row=row, column=1, value=court).font = title_font
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    row += 1

    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    row += 1

    if subtitle:
        ws.cell(row=row, column=1, value=subtitle).font = sub_font
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        row += 1

    ws.cell(row=row, column=1, value=f"Generated: {generated_at}").font = sub_font
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    row += 1

    if filters_summary:
        ws.cell(row=row, column=1, value=f"Filters: {filters_summary}").font = sub_font
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        row += 1

    row += 1  # blank spacer

    def _write_header():
        nonlocal row
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 22
        row += 1

    def _write_row(values):
        nonlocal row
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        row += 1

    total_rows = 0
    if groups:
        for grp in groups:
            ws.cell(row=row, column=1,
                    value=f"{grp['label']} — {grp.get('subtotal', len(grp['rows']))} record(s)"
                    ).font = group_font
            ws.cell(row=row, column=1).fill = group_fill
            ws.merge_cells(f"A{row}:{last_col_letter}{row}")
            row += 1

            _write_header()
            for r in grp["rows"]:
                _write_row(r)

            # subtotal line
            cell = ws.cell(row=row, column=1,
                           value=f"Subtotal: {grp.get('subtotal', len(grp['rows']))}")
            cell.font = subtotal_font
            ws.merge_cells(f"A{row}:{last_col_letter}{row}")
            row += 2  # blank line after each group
            total_rows += grp.get("subtotal", len(grp["rows"]))

        cell = ws.cell(row=row, column=1, value=f"Grand total: {total_rows}")
        cell.font = Font(bold=True, size=11)
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    else:
        _write_header()
        for r in rows:
            _write_row(r)
            total_rows += 1

        cell = ws.cell(row=row, column=1, value=f"Total: {total_rows}")
        cell.font = Font(bold=True, size=11)
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")

    # auto column widths (capped)
    for col_idx, name in enumerate(columns, start=1):
        max_len = len(str(name))
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(60, len(str(cell.value))))
        ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    ws.freeze_panes = "A1"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ===========================================================================
# CSV
# ===========================================================================

def render_csv(
    *,
    columns: Sequence[str],
    rows: Sequence[Sequence],
    filename: str,
    groups: Sequence[dict] | None = None,
) -> HttpResponse:
    """Render a tabular report as a plain UTF-8 CSV download.

    Kept deliberately clean — just a header row followed by data rows — so it
    opens correctly in Excel, LibreOffice, Google Sheets and any importer. A
    UTF-8 BOM is prepended so Excel renders accented characters correctly.
    Grouped reports are flattened (each group prefixed with a label row).
    """
    import csv
    from io import StringIO

    buf = StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    if groups:
        for grp in groups:
            writer.writerow([grp["label"]])
            writer.writerow(list(columns))
            for r in grp["rows"]:
                writer.writerow(["" if v is None else v for v in r])
            writer.writerow([])
    else:
        writer.writerow(list(columns))
        for r in rows:
            writer.writerow(["" if v is None else v for v in r])

    payload = "﻿" + buf.getvalue()  # BOM for Excel
    response = HttpResponse(payload, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ===========================================================================
# Audit helper
# ===========================================================================

def log_report_export(request, report_key: str, fmt: str, applied: dict, row_count: int):
    """Write an EXPORT entry to the audit log. Best-effort — never raises."""
    try:
        from audit.models import AuditLog
        xff = request.META.get("HTTP_X_FORWARDED_FOR")
        ip = xff.split(",")[0].strip() if xff else request.META.get("REMOTE_ADDR")
        AuditLog.objects.create(
            user=getattr(request.user, "username", None) or "anonymous",
            user_email=getattr(request.user, "email", "") or None,
            action="EXPORT",
            model_name="Report",
            record_id="-",
            record_identifier=f"{report_key} ({fmt}, {row_count} rows)",
            new_values={"report": report_key, "format": fmt,
                        "filters": applied, "rows": row_count},
            ip_address=ip,
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
            status="SUCCESS",
        )
    except Exception:
        pass
