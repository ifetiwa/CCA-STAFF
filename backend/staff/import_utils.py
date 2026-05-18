"""
Bulk import utilities for Staff Biodata System.
Handles Excel file processing, validation, and preview.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date, timedelta
from django.core.exceptions import ValidationError
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from staff.models import Staff
from departments.models import Department, Designation, GradeLevel, PostingLocation
from audit.models import AuditLog
import logging
import re

User = get_user_model()
logger = logging.getLogger(__name__)


class StaffImportValidator:
    """Validate staff data for bulk import."""
    
    # Required fields that must be present
    REQUIRED_FIELDS = {
        'staff_id': 'Staff ID',
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'date_of_birth': 'Date of Birth',
        'gender': 'Gender',
        'email': 'Email',
        'phone_number': 'Phone Number',
        'residential_address': 'Residential Address',
        'residential_state': 'Residential State',
        'residential_city': 'Residential City',
        'department': 'Department',
        'designation': 'Designation',
        'employment_type': 'Employment Type',
        'employment_status': 'Employment Status',
        'first_appointment_date': 'First Appointment Date',
    }
    
    # Optional fields
    OPTIONAL_FIELDS = {
        'middle_name': 'Middle Name',
        'nationality': 'Nationality',
        'state_of_origin': 'State of Origin',
        'marital_status': 'Marital Status',
        'last_promotion_date': 'Last Promotion Date',
        'grade_level': 'Grade Level',
        'posting_location': 'Posting Location',
        'next_of_kin_name': 'Next of Kin Name',
        'next_of_kin_phone': 'Next of Kin Phone',
        'bank_name': 'Bank Name',
        'account_number': 'Account Number',
    }
    
    GENDER_CHOICES = {'M', 'F', 'O'}
    EMPLOYMENT_TYPE_CHOICES = {'Permanent', 'Contract', 'Temporary', 'Casual'}
    EMPLOYMENT_STATUS_CHOICES = {
        'Active', 'On Leave', 'Pending', 'Secondment',
        'Retirement', 'Resignation', 'Deceased', 'Archive',
    }
    # Legacy values that older spreadsheets may still contain. Mapped to the
    # new canonical values during import so existing files keep working.
    EMPLOYMENT_STATUS_ALIASES = {
        'Suspended':  'On Leave',
        'Retired':    'Retirement',
        'Terminated': 'Resignation',
    }
    MARITAL_STATUS_CHOICES = {'Single', 'Married', 'Divorced', 'Widowed'}
    
    def __init__(self):
        self.errors = []
        self.warnings = []
    
    def validate_row(self, row_data, row_number):
        """
        Validate a single row of staff data.
        Returns tuple: (is_valid, error_messages, warning_messages)
        """
        row_errors = []
        row_warnings = []
        
        # Check required fields
        for field, label in self.REQUIRED_FIELDS.items():
            value = row_data.get(field, '').strip() if row_data.get(field) else ''
            if not value:
                row_errors.append(f"{label} is required")
        
        if row_errors:
            return False, row_errors, row_warnings
        
        # Validate individual fields
        field_validation = self._validate_fields(row_data, row_number)
        row_errors.extend(field_validation['errors'])
        row_warnings.extend(field_validation['warnings'])
        
        # Check for existing staff ID
        if self._staff_id_exists(row_data.get('staff_id')):
            row_errors.append(f"Staff ID '{row_data.get('staff_id')}' already exists in database")
        
        # Check for existing email
        if self._email_exists(row_data.get('email')):
            row_errors.append(f"Email '{row_data.get('email')}' already exists in database")
        
        is_valid = len(row_errors) == 0
        return is_valid, row_errors, row_warnings
    
    def _validate_fields(self, row_data, row_number):
        """Validate individual field values."""
        errors = []
        warnings = []
        
        # Validate email format
        email = row_data.get('email', '').strip()
        if email and not self._is_valid_email(email):
            errors.append(f"Invalid email format: '{email}'")
        
        # Validate phone number
        phone = row_data.get('phone_number', '').strip()
        if phone and not self._is_valid_phone(phone):
            warnings.append(f"Phone number format may be invalid: '{phone}'")
        
        # Validate gender
        gender = row_data.get('gender', '').strip().upper()
        if gender and gender not in self.GENDER_CHOICES:
            errors.append(f"Gender must be M, F, or O (got: '{gender}')")
        else:
            row_data['gender'] = gender
        
        # Validate employment type
        emp_type = row_data.get('employment_type', '').strip()
        if emp_type and emp_type not in self.EMPLOYMENT_TYPE_CHOICES:
            errors.append(f"Employment Type must be one of {self.EMPLOYMENT_TYPE_CHOICES}")
        
        # Validate employment status. Accept legacy values by rewriting them
        # in-place so the downstream insert uses the canonical name.
        emp_status = row_data.get('employment_status', '').strip()
        if emp_status and emp_status in self.EMPLOYMENT_STATUS_ALIASES:
            row_data['employment_status'] = self.EMPLOYMENT_STATUS_ALIASES[emp_status]
            emp_status = row_data['employment_status']
        if emp_status and emp_status not in self.EMPLOYMENT_STATUS_CHOICES:
            errors.append(
                f"Employment Status must be one of {sorted(self.EMPLOYMENT_STATUS_CHOICES)}"
            )
        
        # Validate marital status
        marital = row_data.get('marital_status', '').strip()
        if marital and marital not in self.MARITAL_STATUS_CHOICES:
            warnings.append(f"Marital Status value may be incorrect: '{marital}'")
        
        # Validate dates
        dob = self._parse_date(row_data.get('date_of_birth'))
        if dob is None and row_data.get('date_of_birth'):
            errors.append(f"Invalid Date of Birth format: '{row_data.get('date_of_birth')}'. Use YYYY-MM-DD or DD/MM/YYYY")
        elif dob:
            row_data['date_of_birth'] = dob
            # Check age is reasonable
            today = date.today()
            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
            if age < 18:
                errors.append(f"Staff member must be at least 18 years old (calculated age: {age})")
            elif age > 80:
                warnings.append(f"Staff member age seems high: {age} years")
        
        # Validate appointment date
        appt_date = self._parse_date(row_data.get('first_appointment_date'))
        if appt_date is None and row_data.get('first_appointment_date'):
            errors.append(f"Invalid First Appointment Date format: '{row_data.get('first_appointment_date')}'. Use YYYY-MM-DD or DD/MM/YYYY")
        elif appt_date:
            row_data['first_appointment_date'] = appt_date
            # Appointment date should be in the past
            if appt_date > date.today():
                errors.append(f"First Appointment Date cannot be in the future")
            # Appointment should be before DOB + 18 years
            if dob and appt_date < (dob + timedelta(days=365*18)):
                errors.append(f"First Appointment Date cannot be before staff member turned 18")
        
        # Validate last promotion date (optional)
        if row_data.get('last_promotion_date'):
            promo_date = self._parse_date(row_data.get('last_promotion_date'))
            if promo_date is None:
                errors.append(f"Invalid Last Promotion Date format: '{row_data.get('last_promotion_date')}'. Use YYYY-MM-DD or DD/MM/YYYY")
            elif appt_date and promo_date < appt_date:
                errors.append(f"Last Promotion Date cannot be before First Appointment Date")
            elif promo_date > date.today():
                errors.append(f"Last Promotion Date cannot be in the future")
            elif promo_date:
                row_data['last_promotion_date'] = promo_date
        
        # Validate foreign keys exist
        dept_name = row_data.get('department', '').strip()
        if dept_name:
            if not Department.objects.filter(name=dept_name).exists():
                errors.append(f"Department '{dept_name}' not found in system")
        
        desig_name = row_data.get('designation', '').strip()
        if desig_name:
            if not Designation.objects.filter(title=desig_name).exists():
                errors.append(f"Designation '{desig_name}' not found in system")
        
        grade_name = row_data.get('grade_level', '').strip()
        if grade_name:
            if not GradeLevel.objects.filter(grade_level=grade_name).exists():
                errors.append(f"Grade Level '{grade_name}' not found in system")
        
        location_name = row_data.get('posting_location', '').strip()
        if location_name:
            if not PostingLocation.objects.filter(name=location_name).exists():
                errors.append(f"Posting Location '{location_name}' not found in system")
        
        return {'errors': errors, 'warnings': warnings}
    
    @staticmethod
    def _is_valid_email(email):
        """Check if email format is valid."""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    @staticmethod
    def _is_valid_phone(phone):
        """Check if phone number format is valid."""
        # Remove common separators
        cleaned = re.sub(r'[\s\-\(\).]', '', phone)
        # Should be at least 10 digits
        return len(cleaned) >= 10 and cleaned.replace('+', '').isdigit()
    
    @staticmethod
    def _parse_date(date_str):
        """
        Parse date string in various formats.
        Supports: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY
        """
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        
        # Try different formats
        formats = [
            '%Y-%m-%d',  # YYYY-MM-DD
            '%d/%m/%Y',  # DD/MM/YYYY
            '%d-%m-%Y',  # DD-MM-YYYY
            '%m/%d/%Y',  # MM/DD/YYYY (US format)
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
    
    @staticmethod
    def _staff_id_exists(staff_id):
        """Check if staff ID already exists."""
        return Staff.objects.filter(staff_id=staff_id.strip()).exists()
    
    @staticmethod
    def _email_exists(email):
        """Check if email already exists."""
        return Staff.objects.filter(email=email.strip().lower()).exists()


class ExcelTemplateGenerator:
    """Generate Excel template for staff import."""
    
    @staticmethod
    def generate_template():
        """Generate Excel template with headers and sample data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Data"
        
        # Define header style
        header_fill = PatternFill(start_color="1a3a52", end_color="1a3a52", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Prepare all field headers
        all_fields = list(StaffImportValidator.REQUIRED_FIELDS.keys()) + \
                    list(StaffImportValidator.OPTIONAL_FIELDS.keys())
        
        field_labels = {
            **StaffImportValidator.REQUIRED_FIELDS,
            **StaffImportValidator.OPTIONAL_FIELDS
        }
        
        # Write headers
        for col_idx, field in enumerate(all_fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field_labels.get(field, field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add sample data row
        sample_data = {
            'staff_id': 'CCL001',
            'first_name': 'John',
            'middle_name': 'Ade',
            'last_name': 'Okafor',
            'date_of_birth': '1985-05-15',
            'gender': 'M',
            'email': 'john.okafor@cca.gov.ng',
            'phone_number': '+234-801-234-5678',
            'residential_address': '123 Lekki Road',
            'residential_state': 'Lagos',
            'residential_city': 'Lagos',
            'nationality': 'Nigerian',
            'state_of_origin': 'Enugu',
            'marital_status': 'Married',
            'department': 'Legal',
            'designation': 'Senior Registrar',
            'employment_type': 'Permanent',
            'employment_status': 'Active',
            'first_appointment_date': '2010-03-01',
            'grade_level': 'GL-12',
            'posting_location': 'Abuja HQ',
        }
        
        for col_idx, field in enumerate(all_fields, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = sample_data.get(field, '')
            cell.border = border
        
        # Set column widths
        column_widths = {
            'staff_id': 12,
            'first_name': 15,
            'middle_name': 15,
            'last_name': 15,
            'date_of_birth': 15,
            'gender': 8,
            'email': 25,
            'phone_number': 15,
            'residential_address': 25,
            'residential_state': 12,
            'residential_city': 12,
            'nationality': 12,
            'state_of_origin': 15,
            'marital_status': 12,
            'department': 15,
            'designation': 20,
            'employment_type': 12,
            'employment_status': 15,
            'first_appointment_date': 15,
            'grade_level': 12,
            'posting_location': 15,
        }
        
        for col_idx, field in enumerate(all_fields, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                column_widths.get(field, 18)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add instructions sheet
        instr_ws = wb.create_sheet("Instructions")
        instructions = [
            ["Staff Import Template Instructions"],
            [],
            ["IMPORTANT NOTES:"],
            ["- All fields marked with * are REQUIRED"],
            ["- Use DD/MM/YYYY or YYYY-MM-DD format for dates"],
            ["- Staff ID must be unique - not already in the system"],
            ["- Email must be unique - not already in the system"],
            ["- Department, Designation, Grade Level, and Posting Location must exist in the system"],
            ["- Do not modify the header row"],
            ["- Maximum 500 rows per import"],
            [],
            ["FIELD DEFINITIONS:"],
            ["Staff ID", "Unique identifier for staff member (e.g., CCL001)"],
            ["First Name", "First name of staff member"],
            ["Last Name", "Last name/surname of staff member"],
            ["Date of Birth", "Date of birth (YYYY-MM-DD)"],
            ["Gender", "M (Male), F (Female), or O (Other)"],
            ["Email", "Official email address"],
            ["Phone Number", "Primary phone number"],
            ["Residential Address", "Current residential address"],
            ["Residential State", "State of residence in Nigeria"],
            ["Residential City", "City of residence"],
            ["Department", "Department name (must exist in system)"],
            ["Designation", "Job designation (must exist in system)"],
            ["Employment Type", "Permanent, Contract, Temporary, or Casual"],
            ["Employment Status", "Active, On Leave, Pending, Secondment, Retirement, Resignation, Deceased, or Archive"],
            ["First Appointment Date", "Date of first appointment (YYYY-MM-DD)"],
            [],
            ["OPTIONAL FIELDS:"],
            ["Grade Level", "Salary grade level (must exist in system)"],
            ["Posting Location", "Posting location (must exist in system)"],
            ["Last Promotion Date", "Date of last promotion (YYYY-MM-DD)"],
            ["Marital Status", "Single, Married, Divorced, or Widowed"],
            ["Next of Kin Name", "Name of next of kin"],
            ["Next of Kin Phone", "Phone number of next of kin"],
            ["Bank Name", "Name of bank for salary payment"],
            ["Account Number", "Bank account number"],
        ]
        
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                instr_ws.cell(row=row_idx, column=col_idx).value = value
        
        return wb


class BulkStaffImporter:
    """Handle bulk import of staff data from Excel."""
    
    def __init__(self, user, request=None):
        self.user = user
        self.request = request
        self.validator = StaffImportValidator()
    
    def parse_excel_file(self, file_path):
        """
        Parse Excel file and return list of staff data.
        Returns: (parsed_rows, errors)
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            parsed_rows = []
            file_errors = []
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
                else:
                    break
            
            if not headers:
                return [], ["No headers found in Excel file"]
            
            # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                # Stop if we hit an empty row
                if not any(cell.value for cell in row[:len(headers)]):
                    break
                
                # Build row data dictionary
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell = row[col_idx]
                    value = cell.value
                    
                    # Handle different cell types
                    if value is None:
                        row_data[header] = ''
                    elif isinstance(value, datetime):
                        row_data[header] = value.date()
                    else:
                        row_data[header] = str(value).strip()
                
                # Add row number for reference
                row_data['_row_number'] = row_idx
                parsed_rows.append(row_data)
            
            if not parsed_rows:
                file_errors.append("No data rows found in Excel file")
            
            if len(parsed_rows) > 500:
                file_errors.append(f"Too many rows ({len(parsed_rows)}). Maximum 500 rows allowed")
            
            return parsed_rows, file_errors
        
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            return [], [f"Error reading Excel file: {str(e)}"]
    
    def validate_and_preview(self, parsed_rows):
        """
        Validate all rows and return preview data with validation results.
        Returns: (valid_rows, invalid_rows, summary)
        """
        valid_rows = []
        invalid_rows = []
        
        for row_data in parsed_rows:
            row_num = row_data.get('_row_number', '?')
            is_valid, errors, warnings = self.validator.validate_row(row_data, row_num)
            
            row_info = {
                'row_number': row_num,
                'data': row_data,
                'is_valid': is_valid,
                'errors': errors,
                'warnings': warnings,
            }
            
            if is_valid:
                valid_rows.append(row_info)
            else:
                invalid_rows.append(row_info)
        
        summary = {
            'total_rows': len(parsed_rows),
            'valid_rows': len(valid_rows),
            'invalid_rows': len(invalid_rows),
        }
        
        return valid_rows, invalid_rows, summary
    
    @transaction.atomic
    def import_valid_rows(self, valid_rows):
        """
        Import all valid rows into the database.
        Returns: (imported_count, error_details, audit_log_id)
        """
        imported_count = 0
        import_errors = []
        created_staff_ids = []
        
        for row_info in valid_rows:
            row_data = row_info['data']
            row_num = row_info['row_number']
            
            try:
                # Prepare data for Staff model
                staff_data = self._prepare_staff_data(row_data)
                
                # Create staff record
                staff = Staff.objects.create(**staff_data)
                imported_count += 1
                created_staff_ids.append(staff.staff_id)
                
                logger.info(f"Imported staff: {staff.staff_id} - {staff.get_full_name()}")
            
            except Exception as e:
                error_msg = f"Row {row_num}: Failed to create staff record - {str(e)}"
                import_errors.append(error_msg)
                logger.error(error_msg)
        
        # Log bulk import action in audit trail
        audit_log_id = self._log_bulk_import(
            imported_count=imported_count,
            failed_count=len(import_errors),
            staff_ids=created_staff_ids,
            errors=import_errors
        )
        
        return imported_count, import_errors, audit_log_id
    
    @staticmethod
    def _prepare_staff_data(row_data):
        """Prepare row data for Staff model creation."""
        # Get foreign key objects
        department = Department.objects.get(name=row_data['department'].strip())
        designation = Designation.objects.get(title=row_data['designation'].strip())
        
        grade_level = None
        if row_data.get('grade_level'):
            try:
                grade_level = GradeLevel.objects.get(grade_level=row_data['grade_level'].strip())
            except GradeLevel.DoesNotExist:
                pass
        
        posting_location = None
        if row_data.get('posting_location'):
            try:
                posting_location = PostingLocation.objects.get(name=row_data['posting_location'].strip())
            except PostingLocation.DoesNotExist:
                pass
        
        # Map row data to Staff fields
        staff_data = {
            'staff_id': row_data['staff_id'].strip(),
            'first_name': row_data['first_name'].strip(),
            'middle_name': row_data.get('middle_name', '').strip() or None,
            'last_name': row_data['last_name'].strip(),
            'date_of_birth': row_data['date_of_birth'],
            'gender': row_data['gender'].strip().upper(),
            'email': row_data['email'].strip().lower(),
            'phone_number': row_data['phone_number'].strip(),
            'residential_address': row_data['residential_address'].strip(),
            'residential_state': row_data['residential_state'].strip(),
            'residential_city': row_data['residential_city'].strip(),
            'nationality': row_data.get('nationality', 'Nigerian').strip() or 'Nigerian',
            'state_of_origin': row_data.get('state_of_origin', '').strip(),
            'marital_status': row_data.get('marital_status', 'Single').strip(),
            'department': department,
            'designation': designation,
            'employment_type': row_data['employment_type'].strip(),
            'employment_status': row_data['employment_status'].strip(),
            'first_appointment_date': row_data['first_appointment_date'],
            'grade_level': grade_level,
            'posting_location': posting_location,
            'is_active': True,
        }
        
        # Add optional fields
        if row_data.get('last_promotion_date'):
            staff_data['last_promotion_date'] = row_data['last_promotion_date']
        
        if row_data.get('next_of_kin_name'):
            staff_data['next_of_kin_name'] = row_data['next_of_kin_name'].strip()
        
        if row_data.get('next_of_kin_phone'):
            staff_data['next_of_kin_phone'] = row_data['next_of_kin_phone'].strip()
        
        if row_data.get('bank_name'):
            staff_data['bank_name'] = row_data['bank_name'].strip()
        
        if row_data.get('account_number'):
            staff_data['account_number'] = row_data['account_number'].strip()
        
        return staff_data
    
    def _log_bulk_import(self, imported_count, failed_count, staff_ids, errors):
        """Log bulk import action in audit trail."""
        try:
            audit_log = AuditLog.objects.create(
                user=self.user,
                action='BULK_IMPORT',
                model_name='Staff',
                record_id=None,  # Bulk action
                old_values=None,
                new_values={
                    'imported_count': imported_count,
                    'failed_count': failed_count,
                    'staff_ids': staff_ids,
                    'error_details': errors[:10]  # First 10 errors
                },
                changed_fields=['bulk_import'],
                ip_address=self._get_client_ip(),
                user_agent=self._get_user_agent(),
                request_method='POST',
                status='SUCCESS' if imported_count > 0 else 'FAILED',
            )
            return audit_log.id
        except Exception as e:
            logger.error(f"Failed to log bulk import: {str(e)}")
            return None
    
    def _get_client_ip(self):
        """Extract client IP from the request, honoring X-Forwarded-For when present."""
        if not self.request:
            return None
        forwarded = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if forwarded:
            return forwarded.split(',')[0].strip()
        return self.request.META.get('REMOTE_ADDR')

    def _get_user_agent(self):
        """Extract user agent from the request."""
        if not self.request:
            return 'Bulk Import Tool'
        return self.request.META.get('HTTP_USER_AGENT', 'Bulk Import Tool')[:255]
